#!/usr/bin/env python3
# ============================================================
# Converte o export do GoodData ("Validação Lançamentos c Histórico")
# para o formato do IMPORTADOR de Orçado Baseline do Planorc
# (menu Importar ▾ → Orçado Baseline, formato largo).
#
# Uso:
#   python3 converter_orcado_baseline.py <origem.xlsx> <saida.xlsx> [--mapa-cc <ref.xlsx>]
#
# ORIGEM (GoodData): linha 1 = rótulos de mês (Jan 2026…Dec 2026) sobre as
#   colunas de valor; linha 2 = cabeçalho (Bk Codorc1, Código da Empresa,
#   Empresa, Código da Filial, Filial, Código Centro de Custo, Desc CC,
#   Bk Codorc3, Desc Orc3, HistóricoOrçamento, Orc_ValorOrc ×12).
# SAÍDA (aba Dados): Empresa, Filial, ItemOrcamento, Centro De Custo,
#   Area, Divisão, BU, Histórico, colunas-data (2026-01-01…12-01) —
#   códigos como TEXTO (preserva zeros à esquerda) + aba Relatorio_Importacao.
# Area/Divisão/BU não existem na origem: vêm do --mapa-cc (qualquer xlsx com
#   aba Dados no mesmo formato, ex.: a conversão anterior; CC → trio mais
#   frequente). CC sem mapa sai '0' e é listado no log.
# ============================================================
import sys, re, datetime
from collections import Counter
from openpyxl import load_workbook, Workbook

MES_EN = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def txt(v):
    return ('' if v is None else str(v)).strip()

def carregar_mapa_cc(path):
    mapa = {}
    if not path: return mapa
    wb = load_workbook(path, read_only=True)
    ws = wb['Dados'] if 'Dados' in wb.sheetnames else wb.active
    cont = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        cc = txt(r[3])
        if not cc: continue
        trio = (txt(r[4]), txt(r[5]), txt(r[6]))
        cont.setdefault(cc, Counter())[trio] += 1
    for cc, c in cont.items(): mapa[cc] = c.most_common(1)[0][0]
    return mapa

def main():
    args = sys.argv[1:]
    if len(args) < 2:
        print('uso: converter_orcado_baseline.py <origem.xlsx> <saida.xlsx> [--mapa-cc ref.xlsx]'); sys.exit(1)
    src, out = args[0], args[1]
    mapa_path = args[args.index('--mapa-cc') + 1] if '--mapa-cc' in args else None
    mapa = carregar_mapa_cc(mapa_path)

    wb = load_workbook(src, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    rotulos = next(it)   # linha 1: rótulos de mês sobre as colunas de valor
    header  = next(it)   # linha 2: cabeçalho real
    # localizar colunas de valor pelo rótulo "Mes AAAA" da linha 1
    meses = []           # [(col_idx, ano, mes)]
    for i, v in enumerate(rotulos):
        m = re.match(r'^([A-Za-z]{3})\w*[ /-](\d{4})$', txt(v))
        if m and m.group(1).lower()[:3] in MES_EN:
            meses.append((i, int(m.group(2)), MES_EN[m.group(1).lower()[:3]]))
    if not meses: raise SystemExit('Não achei os rótulos de mês na linha 1 (ex.: "Jan 2026")')
    anos = sorted({a for _, a, _ in meses})
    if len(anos) > 1: raise SystemExit(f'Origem cruza anos {anos} — o importador espera 1 ano por arquivo')
    ano = anos[0]
    # índices das dimensões pelo cabeçalho da linha 2 (busca por conteúdo)
    def col(*pats):
        for i, h in enumerate(header):
            t = txt(h).lower()
            if t and all(p in t for p in pats): return i
        raise SystemExit(f'Coluna não encontrada no cabeçalho: {pats}')
    c_emp, c_fil = col('código', 'empresa'), col('código', 'filial')
    c_cc, c_item = col('código', 'centro'), col('codorc3')
    c_hist = col('histórico')

    out_wb = Workbook(write_only=True)
    ds = out_wb.create_sheet('Dados')
    ds.append(['Empresa', 'Filial', 'ItemOrcamento', 'Centro De Custo', 'Area', 'Divisão', 'BU', 'Histórico']
              + [datetime.datetime(ano, m, 1) for _, _, m in meses])
    n, ign, tot = 0, 0, Counter()
    cc_sem_mapa = set()
    for r in it:
        emp, fil = txt(r[c_emp]), txt(r[c_fil])
        item, cc = txt(r[c_item]), txt(r[c_cc])
        if not emp or not item:
            ign += 1; continue
        area, div, bu = mapa.get(cc, ('0', '0', '0'))
        if cc and cc not in mapa: cc_sem_mapa.add(cc)
        vals = []
        for ci, _, m in meses:
            v = r[ci]
            v = float(str(v).replace(',', '.')) if v not in (None, '') else 0.0
            vals.append(v); tot[m] += v
        ds.append([emp, fil, item, cc, area or '0', div or '0', bu or '0', txt(r[c_hist])] + vals)
        n += 1
    log = out_wb.create_sheet('Relatorio_Importacao')
    log.append(['Arquivo', 'Aba', 'Linhas convertidas', 'Ignoradas (sem empresa/item)', 'Ano', 'CCs sem mapa'])
    log.append([src.split('/')[-1], ws.title, n, ign, ano, ', '.join(sorted(cc_sem_mapa)) or '—'])
    out_wb.save(out)
    print(f'{n} linhas convertidas ({ign} ignoradas) · ano {ano} · CCs sem mapa: {sorted(cc_sem_mapa) or "nenhum"}')
    for m in sorted(tot): print(f'  {ano}-{m:02d}: {tot[m]:,.2f}')

if __name__ == '__main__':
    main()
